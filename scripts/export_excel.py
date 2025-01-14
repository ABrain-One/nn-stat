from openpyxl import Workbook
from openpyxl.drawing.image import Image
import os


def export_to_excel(aggregated_data, output_file, plot_dir, image_width=400, image_height=300):
    """
    Export aggregated data and categorized plots to an Excel file.
    Args:
        aggregated_data (DataFrame): Aggregated statistics.
        raw_data (DataFrame): Raw data.
        output_file (str): Path to save the Excel file.
        plot_dir (str): Directory containing plots.
        image_width (int): Width of embedded images (default: 400).
        image_height (int): Height of embedded images (default: 300).
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Statistics"

    # Add headers
    ws.append(aggregated_data.columns.tolist())

    # Add data rows
    for row in aggregated_data.itertuples(index=False):
        ws.append(row)

    # Define plot categories and their corresponding file patterns
    parallel_categories = {
        "Mean and Std Dev": "{task}_{dataset}_{metric}_mean_std.png",
        "Box Plot": "{task}_{dataset}_{metric}_box.png",
        "Rolling Mean": "{task}_{dataset}_{metric}_rolling_mean.png",
    }
    sequential_categories = {
        "Task Comparison (Accuracy)": "task_comparison_accuracy.png",
        "Task Comparison (IoU)": "task_comparison_iou.png",
        "Correlation Heatmap": "correlation_heatmap.png",
        "Time vs Accuracy": "time_vs_accuracy.png",
        "Time vs IoU": "time_vs_iou.png",
    }

    # Add parallel layout plots (adjust positions for each category)
    start_row = 2  # Start embedding images below the data headers
    rightmost_column = len(aggregated_data.columns) + 3  # Move two blocks left

    for category, file_suffix in parallel_categories.items():
        if category == "Mean and Std Dev":
            # Display "Mean and Std Dev" at its current position
            category_row = start_row - 1
        elif category == "Box Plot":
            # Shift "Box Plot" three blocks down
            start_row += 3
            category_row = start_row - 1
        elif category == "Rolling Mean":
            # Shift "Rolling Mean" three blocks further down
            start_row += 3
            category_row = start_row - 1

        # Add the category name above the row of images
        ws.cell(row=category_row, column=rightmost_column, value=category)
        col_offset = rightmost_column  # Start placing images

        for metric in aggregated_data['metric'].unique():
            for (task, dataset) in aggregated_data.groupby(['task', 'dataset']).groups.keys():
                plot_file = file_suffix.format(task=task, dataset=dataset, metric=metric).replace(" ", "_")
                plot_path = os.path.join(plot_dir, plot_file)
                if os.path.exists(plot_path):
                    img = Image(plot_path)
                    img.width = image_width
                    img.height = image_height
                    img.anchor = f"{chr(64 + col_offset)}{start_row}"  # Place image in a new column
                    ws.add_image(img)
                    col_offset += 6  # Adjust column spacing for images

        start_row += 15  # Move to the next row for the next category

    # Add sequential layout plots
    start_row += 5  # Leave extra space after parallel layout
    for category, file_suffix in sequential_categories.items():
        if category == "Task Comparison (Accuracy)":
            start_row += 8  # Move the name 8 blocks down

        ws.merge_cells(start_row=start_row, start_column=1, end_row=start_row, end_column=8)
        ws.cell(row=start_row, column=1, value=category)
        start_row += 3  # Leave a gap for better readability

        plot_path = os.path.join(plot_dir, file_suffix)
        if os.path.exists(plot_path):
            img = Image(plot_path)
            img.width = image_width
            img.height = image_height
            img.anchor = f"A{start_row}"  # Place image vertically
            ws.add_image(img)
            start_row += 22  # Adjust row spacing for resized images

    # Save the Excel file
    wb.save(output_file)
    print(f"Excel file saved to {output_file}")
