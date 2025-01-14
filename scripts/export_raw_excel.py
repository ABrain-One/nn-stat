from openpyxl.utils import get_column_letter
import pandas as pd
import matplotlib.pyplot as plt
from openpyxl import Workbook
from openpyxl.drawing.image import Image
import os



def create_plot(data, x_column, y_column, plot_type, title, output_path):
    """
    Create different types of plots and save them as image files.
    Args:
        data (DataFrame): Data for plotting.
        x_column (str): Column name for x-axis.
        y_column (str): Column name for y-axis.
        plot_type (str): Type of plot (scatter, line, box, histogram).
        title (str): Plot title.
        output_path (str): Path to save the plot.
    """
    plt.figure(figsize=(8, 6))

    if plot_type == "scatter":
        plt.scatter(data[x_column], data[y_column], alpha=0.7, edgecolor='k')
    elif plot_type == "line":
        plt.plot(data[x_column], data[y_column], marker='o', linestyle='-', alpha=0.7)
    elif plot_type == "box":
        data.boxplot(column=y_column, by=x_column)
        plt.title(title)
        plt.suptitle("")  # Remove default subtitle
    elif plot_type == "histogram":
        plt.hist(data[y_column], bins=20, alpha=0.7, edgecolor='k')

    plt.title(title)
    plt.xlabel(x_column.capitalize())
    plt.ylabel(y_column.capitalize())
    plt.grid(True)
    plt.savefig(output_path)
    plt.close()


def export_raw_data_with_plots(filtered_raw_data, output_file, plot_dir):
    """
    Export raw data and plots to an Excel file, positioning the plots vertically.
    Args:
        filtered_raw_data (DataFrame): Filtered raw data.
        output_file (str): Path to save the Excel file.
        plot_dir (str): Directory to save plots.
    """
    os.makedirs(plot_dir, exist_ok=True)

    # Define plots to create
    plot_configs = [
        ("epoch", "accuracy", "scatter", "Accuracy vs Epochs", "accuracy_vs_epochs_scatter.png"),
        ("duration", "accuracy", "scatter", "Accuracy vs Training Duration", "accuracy_vs_duration_scatter.png"),
        ("epoch", "accuracy", "line", "Accuracy Trend over Epochs", "accuracy_vs_epochs_line.png"),
        ("duration", "accuracy", "line", "Accuracy Trend over Training Duration", "accuracy_vs_duration_line.png"),
        ("epoch", "accuracy", "box", "Accuracy Distribution by Epochs", "accuracy_vs_epochs_box.png"),
        ("duration", "accuracy", "box", "Accuracy Distribution by Training Duration", "accuracy_vs_duration_box.png"),
        ("accuracy", "accuracy", "histogram", "Accuracy Frequency", "accuracy_histogram.png"),
    ]

    # Generate plots
    plot_paths = []
    for x, y, plot_type, title, filename in plot_configs:
        output_path = os.path.join(plot_dir, filename)
        create_plot(filtered_raw_data, x, y, plot_type, title, output_path)
        plot_paths.append((title, output_path))

    # Export data and plots to Excel
    wb = Workbook()
    ws = wb.active
    ws.title = "Raw Data"

    # Add headers
    ws.append(filtered_raw_data.columns.tolist())

    # Add data rows
    for row in filtered_raw_data.itertuples(index=False):
        ws.append(row)

    # Embed plots vertically in Excel
    start_row = 2  # Start embedding images after headers
    plot_column_start = len(filtered_raw_data.columns) + 5  # Leave 5 columns after the data

    for title, path in plot_paths:
        if os.path.exists(path):
            # Add plot title
            ws.cell(row=start_row, column=plot_column_start, value=title)
            img = Image(path)
            img.width = 400
            img.height = 300
            img.anchor = f"{get_column_letter(plot_column_start)}{start_row + 1}"  # Place below the title
            ws.add_image(img)
            start_row += 20  # Move to the next row for the next plot

    # Save Excel file
    wb.save(output_file)
    print(f"Raw data and plots saved to {output_file}")
